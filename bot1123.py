import json
import logging
import os
import shutil
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes
import openpyxl

# ============================================
#  НАСТРОЙКИ
# ============================================

BOT_TOKEN = "8239333710:AAHtiMFQSNoc67WWCcHezW7m-hQ9FjxjfZ8"

# Пути к файлам
SCHEDULE_JSON = 'schedule.json'
FILE_1 = '7d5387db9491af08bdc7e1738e02d263.xlsx'
FILE_2 = 'ae2454c52a6854c7272c28bf32f8d521.xlsx'
ADMIN_FILE = 'admin_id.txt'

# ============================================
#  ДНИ НЕДЕЛИ
# ============================================

DAYS_RU = {
    "Понедельник": 0,
    "Вторник": 1,
    "Среда": 2,
    "Четверг": 3,
    "Пятница": 4,
    "Суббота": 5,
    "Воскресенье": 6
}

DAYS_RU_REVERSE = {v: k for k, v in DAYS_RU.items()}

DAYS_SHORT = {
    "Понедельник": "ПН",
    "Вторник": "ВТ",
    "Среда": "СР",
    "Четверг": "ЧТ",
    "Пятница": "ПТ",
    "Суббота": "СБ"
}

# ============================================
#  ЛОГИРОВАНИЕ
# ============================================

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ============================================
#  РАБОТА С РАСПИСАНИЕМ
# ============================================

def parse_schedule(file_path, week_number):
    """
    Парсит Excel-файл и возвращает словарь с расписанием.
    
    Структура:
    {
        "Группа1": {
            "Понедельник": {
                "08:00 - 09:20": ["Предмет1", "Предмет2"],
                ...
            },
            ...
        },
        ...
    }
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
    except Exception as e:
        logger.error(f"Ошибка при загрузке файла {file_path}: {e}")
        return {}

    schedule = {}
    groups_row = 4  # 5-я строка (индексация с 0)

    # --- 1. Собираем все группы из заголовков ---
    group_columns = {}
    for col_idx, cell in enumerate(sheet[groups_row]):
        if col_idx < 2:  # Пропускаем столбцы A (День) и B (Урок)
            continue
        
        group_name = cell.value
        if not group_name:
            continue
            
        # Проверяем, что это группа (содержит характерные символы)
        if any(c in group_name for c in ['Э-', 'ГМУ', 'М-', 'ТД', 'ЭБ']):
            # Обработка групп через ";"
            if ';' in group_name:
                for g in group_name.split(';'):
                    clean_name = g.strip()
                    if clean_name:
                        group_columns[col_idx] = clean_name
                        schedule[clean_name] = {}
            else:
                group_columns[col_idx] = group_name
                schedule[group_name] = {}

    if not group_columns:
        logger.warning(f"В файле {file_path} не найдено ни одной группы")
        return {}

    # --- 2. Парсим расписание по дням и времени ---
    current_day = None
    for row_idx in range(5, sheet.max_row):
        day_cell = sheet.cell(row=row_idx, column=1).value  # Столбец A
        time_cell = sheet.cell(row=row_idx, column=2).value  # Столбец B

        # Если нашли новый день
        if day_cell and day_cell in DAYS_RU:
            current_day = day_cell
            # Инициализируем день для всех групп
            for group in schedule.keys():
                if current_day not in schedule[group]:
                    schedule[group][current_day] = {}

        # Если есть время и текущий день определен
        if time_cell and current_day and isinstance(time_cell, str) and ':' in time_cell:
            for col_idx, group_name in group_columns.items():
                cell_value = sheet.cell(row=row_idx, column=col_idx + 1).value
                if cell_value:
                    info = str(cell_value).strip()
                    # Пропускаем пустые или служебные значения
                    if info and info not in ['', 'None', 'NULL']:
                        if time_cell not in schedule[group_name][current_day]:
                            schedule[group_name][current_day][time_cell] = []
                        schedule[group_name][current_day][time_cell].append(info)

    return schedule

def update_schedule():
    """
    Обновляет JSON-файл с расписанием из Excel-файлов.
    Возвращает True в случае успеха.
    """
    try:
        logger.info("Начинаю обновление расписания...")
        
        # Проверяем наличие файлов
        if not os.path.exists(FILE_1):
            logger.error(f"Файл 1-й недели не найден: {FILE_1}")
            return False
        
        if not os.path.exists(FILE_2):
            logger.error(f"Файл 2-й недели не найден: {FILE_2}")
            return False
        
        # Парсим обе недели
        week1 = parse_schedule(FILE_1, 1)
        week2 = parse_schedule(FILE_2, 2)
        
        if not week1:
            logger.error("Не удалось распарсить 1-ю неделю")
            return False
        
        if not week2:
            logger.error("Не удалось распарсить 2-ю неделю")
            return False
        
        # Объединяем
        full_schedule = {
            "1": week1,
            "2": week2
        }
        
        # Сохраняем в JSON
        with open(SCHEDULE_JSON, 'w', encoding='utf-8') as f:
            json.dump(full_schedule, f, ensure_ascii=False, indent=4)
        
        logger.info("Расписание успешно обновлено!")
        return True
        
    except Exception as e:
        logger.error(f"Ошибка при обновлении расписания: {e}")
        return False

def load_schedule():
    """
    Загружает расписание из JSON-файла.
    Если JSON отсутствует, создает его из Excel.
    """
    global SCHEDULE_DATA
    
    try:
        if os.path.exists(SCHEDULE_JSON):
            with open(SCHEDULE_JSON, 'r', encoding='utf-8') as f:
                SCHEDULE_DATA = json.load(f)
            logger.info("Расписание загружено из JSON")
            
            # Проверяем структуру данных
            if "1" not in SCHEDULE_DATA or "2" not in SCHEDULE_DATA:
                logger.warning("Неверная структура JSON, пересоздаю...")
                raise ValueError("Invalid schedule structure")
                
        else:
            logger.warning("JSON-файл не найден, создаю из Excel...")
            if update_schedule():
                with open(SCHEDULE_JSON, 'r', encoding='utf-8') as f:
                    SCHEDULE_DATA = json.load(f)
            else:
                SCHEDULE_DATA = {"1": {}, "2": {}}
                logger.error("Не удалось загрузить расписание")
                
    except Exception as e:
        logger.error(f"Ошибка при загрузке расписания: {e}")
        SCHEDULE_DATA = {"1": {}, "2": {}}

# Глобальная переменная для расписания
SCHEDULE_DATA = {}

# ============================================
#  ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ============================================

def get_all_groups():
    """Возвращает отсортированный список всех групп."""
    groups = set()
    for week in ['1', '2']:
        for group in SCHEDULE_DATA.get(week, {}).keys():
            if group:
                groups.add(group)
    return sorted(list(groups))

def get_week_number():
    """
    Определяет номер недели (1 или 2) на основе текущей даты.
    !!! ВАЖНО: Установите реальную дату начала семестра !!!
    """
    # ДАТА НАЧАЛА СЕМЕСТРА — ИЗМЕНИТЕ ПОД СЕБЯ!
    start_date = datetime(2024, 9, 1)
    
    today = datetime.now()
    delta = today - start_date
    
    # Если семестр еще не начался
    if delta.days < 0:
        return "1"
    
    # Номер недели (1 или 2)
    week_num = (delta.days // 7) % 2 + 1
    return str(week_num)

def get_schedule_for_day(group_name, day_name, week_num=None):
    """
    Возвращает отформатированное расписание для группы на конкретный день.
    """
    if week_num is None:
        week_num = get_week_number()
    
    # Проверяем существование группы
    if group_name not in SCHEDULE_DATA.get(week_num, {}):
        return None
    
    week_schedule = SCHEDULE_DATA[week_num][group_name]
    
    # Проверяем существование дня
    if day_name not in week_schedule:
        return f"📭 На *{day_name}* занятий нет."
    
    day_schedule = week_schedule[day_name]
    
    # Формируем ответ
    response = f"📚 *{day_name} ({week_num} неделя)*\n"
    response += f"Группа: *{group_name}*\n\n"
    
    # Сортируем по времени
    sorted_times = sorted(day_schedule.keys())
    
    for time_slot in sorted_times:
        response += f"🕐 *{time_slot}*\n"
        for item in day_schedule[time_slot]:
            # Убираем лишние пробелы и переносы
            clean_item = ' '.join(item.split())
            response += f"   • {clean_item}\n"
        response += "\n"
    
    return response

def get_week_schedule(group_name):
    """
    Возвращает отформатированное расписание на всю неделю.
    """
    week_num = get_week_number()
    
    if group_name not in SCHEDULE_DATA.get(week_num, {}):
        return None
    
    week_schedule = SCHEDULE_DATA[week_num][group_name]
    
    response = f"📅 *Расписание на неделю ({week_num} неделя)*\n"
    response += f"Группа: *{group_name}*\n\n"
    
    for day in DAYS_RU.keys():
        if day in week_schedule:
            response += f"*{day}*\n"
            day_schedule = week_schedule[day]
            sorted_times = sorted(day_schedule.keys())
            for time_slot in sorted_times:
                for item in day_schedule[time_slot]:
                    clean_item = ' '.join(item.split())
                    response += f"   {time_slot} — {clean_item}\n"
            response += "\n"
        else:
            response += f"*{day}* — выходной\n\n"
    
    return response

# ============================================
#  КЛАВИАТУРЫ
# ============================================

def get_groups_keyboard():
    """Создает клавиатуру со списком групп (по 2 в ряд)."""
    groups = get_all_groups()
    
    if not groups:
        keyboard = [[InlineKeyboardButton("⏳ Расписание загружается...", callback_data="noop")]]
        return InlineKeyboardMarkup(keyboard)
    
    keyboard = []
    row = []
    for i, group in enumerate(groups):
        row.append(InlineKeyboardButton(group, callback_data=f"group_{group}"))
        if len(row) == 2:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    
    return InlineKeyboardMarkup(keyboard)

def get_day_keyboard(group_name):
    """Создает клавиатуру с основными действиями для выбранной группы."""
    keyboard = [
        [
            InlineKeyboardButton("📖 Сегодня", callback_data=f"day_today_{group_name}"),
            InlineKeyboardButton("📖 Завтра", callback_data=f"day_tomorrow_{group_name}")
        ],
        [
            InlineKeyboardButton("📅 Вся неделя", callback_data=f"week_{group_name}")
        ],
        [
            InlineKeyboardButton("📋 День недели", callback_data=f"show_days_{group_name}")
        ],
        [
            InlineKeyboardButton("🔙 Назад к группам", callback_data="back_to_groups")
        ]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_day_buttons(group_name):
    """Создает клавиатуру с кнопками для выбора конкретного дня."""
    keyboard = []
    row = []
    for day_ru, day_short in DAYS_SHORT.items():
        row.append(InlineKeyboardButton(day_short, callback_data=f"day_{day_ru}_{group_name}"))
        if len(row) == 3:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    
    keyboard.append([InlineKeyboardButton("🔙 Назад", callback_data=f"back_to_group_{group_name}")])
    return InlineKeyboardMarkup(keyboard)

# ============================================
#  ОБРАБОТЧИКИ КОМАНД
# ============================================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start."""
    user_id = update.effective_user.id
    
    # Назначаем администратора, если его еще нет
    admin_id = get_admin_id()
    if admin_id is None:
        set_admin_id(user_id)
        await update.message.reply_text(
            f"✅ Вы назначены администратором бота!\n"
            f"Ваш ID: `{user_id}`\n\n"
            f"Теперь вы можете использовать команду /update_schedule для обновления расписания.",
            parse_mode="Markdown"
        )
    
    # Приветствие
    await update.message.reply_text(
        "👋 *Привет! Я бот расписания экономического факультета.*\n\n"
        "📌 *Как я работаю:*\n"
        "1️⃣ Выбери свою группу из списка ниже\n"
        "2️⃣ Я покажу расписание на *сегодня*\n"
        "3️⃣ Используй кнопки для навигации\n\n"
        "➡️ *Выберите группу:*",
        parse_mode="Markdown",
        reply_markup=get_groups_keyboard()
    )

async def update_schedule_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Команда для администратора: принудительно обновить расписание."""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("⛔ У вас нет прав для этой команды.")
        return
    
    await update.message.reply_text("🔄 Обновляю расписание из Excel-файлов...")
    
    if update_schedule():
        load_schedule()
        await update.message.reply_text(
            "✅ *Расписание успешно обновлено!*\n"
            "Теперь можно пользоваться актуальным расписанием.",
            parse_mode="Markdown",
            reply_markup=get_groups_keyboard()
        )
    else:
        await update.message.reply_text(
            "❌ *Не удалось обновить расписание.*\n\n"
            "Проверьте:\n"
            "• наличие Excel-файлов в папке с ботом\n"
            "• правильность названий файлов\n"
            "• формат файлов (.xlsx)",
            parse_mode="Markdown"
        )

async def show_my_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Команда /myid — показывает ID пользователя."""
    user_id = update.effective_user.id
    username = update.effective_user.username or "не установлен"
    first_name = update.effective_user.first_name or ""
    
    is_admin_text = "👑 *Вы администратор!*" if is_admin(user_id) else "👤 Обычный пользователь"
    
    await update.message.reply_text(
        f"🆔 *Ваш ID в Telegram:*\n"
        f"`{user_id}`\n\n"
        f"👤 Имя: {first_name}\n"
        f"📛 Username: @{username}\n"
        f"Статус: {is_admin_text}\n\n"
        f"Если вы администратор, этот ID можно использовать для настройки.",
        parse_mode="Markdown"
    )

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Команда /help — показывает справку."""
    help_text = (
        "📖 *Как пользоваться ботом*\n\n"
        "🔹 *Выбор группы*\n"
        "   Нажми на свою группу в списке — я покажу расписание на сегодня.\n\n"
        "🔹 *Навигация*\n"
        "   • *Сегодня* — расписание на текущий день\n"
        "   • *Завтра* — расписание на завтра\n"
        "   • *Вся неделя* — полное расписание на неделю\n"
        "   • *День недели* — выбрать конкретный день\n\n"
        "🔹 *Команды*\n"
        "   /start — начать работу\n"
        "   /help — эта справка\n"
        "   /myid — показать мой ID\n"
        "   /update_schedule — обновить расписание (админ)\n\n"
        "📌 *Бот автоматически определяет 1-ю или 2-ю неделю!*"
    )
    
    await update.message.reply_text(help_text, parse_mode="Markdown")

# ============================================
#  ФУНКЦИИ ДЛЯ РАБОТЫ С АДМИНИСТРАТОРОМ
# ============================================

def get_admin_id():
    """Возвращает ID администратора из файла."""
    try:
        if os.path.exists(ADMIN_FILE):
            with open(ADMIN_FILE, 'r') as f:
                return int(f.read().strip())
    except Exception as e:
        logger.error(f"Ошибка при чтении admin_id: {e}")
    return None

def set_admin_id(user_id):
    """Сохраняет ID администратора в файл."""
    try:
        with open(ADMIN_FILE, 'w') as f:
            f.write(str(user_id))
        logger.info(f"Администратор установлен: {user_id}")
        return True
    except Exception as e:
        logger.error(f"Ошибка при сохранении admin_id: {e}")
        return False

def is_admin(user_id):
    """Проверяет, является ли пользователь администратором."""
    admin_id = get_admin_id()
    return admin_id is not None and user_id == admin_id

# ============================================
#  ОБРАБОТЧИКИ КНОПОК
# ============================================

async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик нажатий на инлайн-кнопки."""
    query = update.callback_query
    await query.answer()
    
    data = query.data
    
    # --- Служебные ---
    if data == "noop":
        await query.edit_message_text(
            "⏳ Расписание еще загружается... Попробуйте позже.",
            reply_markup=get_groups_keyboard()
        )
        return
    
    if data == "back_to_groups":
        await query.edit_message_text(
            "👥 *Выберите группу:*",
            parse_mode="Markdown",
            reply_markup=get_groups_keyboard()
        )
        return
    
    # --- Выбор группы ---
    if data.startswith("group_"):
        group_name = data[6:]
        context.user_data['current_group'] = group_name
        
        today_weekday = datetime.now().weekday()
        today_name = DAYS_RU_REVERSE.get(today_weekday)
        
        if today_name:
            schedule = get_schedule_for_day(group_name, today_name)
            if schedule:
                await query.edit_message_text(
                    schedule,
                    parse_mode="Markdown",
                    reply_markup=get_day_keyboard(group_name)
                )
            else:
                await query.edit_message_text(
                    f"❌ Группа '{group_name}' не найдена.\n"
                    "Пожалуйста, выберите другую группу.",
                    parse_mode="Markdown",
                    reply_markup=get_groups_keyboard()
                )
        return
    
    # --- Сегодня ---
    if data.startswith("day_today_"):
        group_name = data[10:]
        today_weekday = datetime.now().weekday()
        today_name = DAYS_RU_REVERSE.get(today_weekday)
        schedule = get_schedule_for_day(group_name, today_name)
        if schedule:
            await query.edit_message_text(
                schedule,
                parse_mode="Markdown",
                reply_markup=get_day_keyboard(group_name)
            )
        return
    
    # --- Завтра ---
    if data.startswith("day_tomorrow_"):
        group_name = data[13:]
        tomorrow = datetime.now() + timedelta(days=1)
        tomorrow_weekday = tomorrow.weekday()
        tomorrow_name = DAYS_RU_REVERSE.get(tomorrow_weekday)
        schedule = get_schedule_for_day(group_name, tomorrow_name)
        if schedule:
            await query.edit_message_text(
                schedule,
                parse_mode="Markdown",
                reply_markup=get_day_keyboard(group_name)
            )
        return
    
    # --- Показать кнопки с днями ---
    if data.startswith("show_days_"):
        group_name = data[10:]
        await query.edit_message_text(
            f"📋 *Выберите день недели для группы {group_name}:*",
            parse_mode="Markdown",
            reply_markup=get_day_buttons(group_name)
        )
        return
    
    # --- Конкретный день ---
    if data.startswith("day_") and not data.startswith("day_today") and not data.startswith("day_tomorrow") and not data.startswith("show_days_"):
        parts = data.split("_")
        if len(parts) >= 3:
            day_name = parts[1]
            group_name = "_".join(parts[2:])
            schedule = get_schedule_for_day(group_name, day_name)
            if schedule:
                await query.edit_message_text(
                    schedule,
                    parse_mode="Markdown",
                    reply_markup=get_day_buttons(group_name)
                )
        return
    
    # --- Вся неделя ---
    if data.startswith("week_"):
        group_name = data[5:]
        schedule = get_week_schedule(group_name)
        if schedule:
            await query.edit_message_text(
                schedule,
                parse_mode="Markdown",
                reply_markup=get_day_keyboard(group_name)
            )
        return
    
    # --- Назад к группе ---
    if data.startswith("back_to_group_"):
        group_name = data[14:]
        today_weekday = datetime.now().weekday()
        today_name = DAYS_RU_REVERSE.get(today_weekday)
        schedule = get_schedule_for_day(group_name, today_name)
        if schedule:
            await query.edit_message_text(
                schedule,
                parse_mode="Markdown",
                reply_markup=get_day_keyboard(group_name)
            )
        return

# ============================================
#  ОБРАБОТЧИК ЗАГРУЗКИ ФАЙЛОВ
# ============================================

async def handle_file_upload(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик загрузки Excel-файлов (только для администратора)."""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("⛔ Вы не можете загружать файлы.")
        return
    
    document = update.message.document
    if not document:
        return
    
    file_name = document.file_name
    if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
        await update.message.reply_text(
            "⚠️ Пожалуйста, загрузите файл Excel (.xlsx или .xls)"
        )
        return
    
    # Скачиваем файл
    file_path = f"temp_{file_name}"
    file = await document.get_file()
    await file.download_to_drive(file_path)
    
    # Определяем, для какой недели файл
    file_lower = file_name.lower()
    if any(x in file_lower for x in ['1 неделя', '1-я', 'week_1', '1_week']):
        target_file = FILE_1
        week_label = "1-й"
    elif any(x in file_lower for x in ['2 неделя', '2-я', 'week_2', '2_week']):
        target_file = FILE_2
        week_label = "2-й"
    else:
        await update.message.reply_text(
            "📁 Не могу определить, для какой недели этот файл.\n\n"
            "Пожалуйста, переименуйте файл, указав в названии:\n"
            "• '1 неделя' или 'week_1' — для 1-й недели\n"
            "• '2 неделя' или 'week_2' — для 2-й недели"
        )
        os.remove(file_path)
        return
    
    # Заменяем файл
    shutil.move(file_path, target_file)
    await update.message.reply_text(f"✅ Файл '{file_name}' сохранен как '{target_file}'")
    
    # Обновляем расписание
    await update.message.reply_text(f"🔄 Обновляю расписание ({week_label} неделя)...")
    
    if update_schedule():
        load_schedule()
        await update.message.reply_text(
            "✅ *Расписание успешно обновлено!*",
            parse_mode="Markdown",
            reply_markup=get_groups_keyboard()
        )
    else:
        await update.message.reply_text(
            "❌ *Ошибка при обновлении расписания.*\n"
            "Проверьте формат файла.",
            parse_mode="Markdown"
        )

# ============================================
#  ОБРАБОТЧИК ТЕКСТОВЫХ СООБЩЕНИЙ
# ============================================

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик текстовых сообщений (если пользователь ввел группу вручную)."""
    group_name = update.message.text.strip().upper()
    
    all_groups = get_all_groups()
    
    # Проверяем точное совпадение
    if group_name in all_groups:
        today_weekday = datetime.now().weekday()
        today_name = DAYS_RU_REVERSE.get(today_weekday)
        schedule = get_schedule_for_day(group_name, today_name)
        if schedule:
            await update.message.reply_text(
                schedule,
                parse_mode="Markdown",
                reply_markup=get_day_keyboard(group_name)
            )
            return
    
    # Если не нашли — показываем список групп
    await update.message.reply_text(
        f"❌ Группа '{group_name}' не найдена.\n\n"
        "Пожалуйста, выберите группу из списка ниже:",
        reply_markup=get_groups_keyboard()
    )

# ============================================
#  ЗАПУСК БОТА
# ============================================

def main():
    """Запуск бота."""
    # Загружаем расписание
    load_schedule()
    
    # Создаем приложение
    application = Application.builder().token(BOT_TOKEN).build()
    
    # Регистрируем обработчики команд
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("myid", show_my_id))
    application.add_handler(CommandHandler("update_schedule", update_schedule_command))
    
    # Регистрируем обработчики кнопок и сообщений
    application.add_handler(CallbackQueryHandler(button_callback))
    application.add_handler(MessageHandler(filters.Document.ALL, handle_file_upload))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    
    # Выводим информацию при запуске
    admin_id = get_admin_id()
    groups = get_all_groups()
    
    print("=" * 50)
    print("🤖 БОТ РАСПИСАНИЯ ЗАПУЩЕН")
    print("=" * 50)
    print(f"👤 Администратор: {'установлен (ID: ' + str(admin_id) + ')' if admin_id else 'не назначен'}")
    print(f"📅 Расписание: {'загружено' if SCHEDULE_DATA else 'не загружено'}")
    print(f"📊 Найдено групп: {len(groups)}")
    print(f"📁 Файлы: {FILE_1}, {FILE_2}")
    print("-" * 50)
    print("Доступные команды:")
    print("  /start            - Начать работу")
    print("  /help             - Показать справку")
    print("  /myid             - Показать мой ID")
    print("  /update_schedule  - Обновить расписание (админ)")
    print("=" * 50)
    
    # Запускаем бота
    application.run_polling()

if __name__ == '__main__':
    main()