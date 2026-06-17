import json
import logging
import os
import shutil
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes
import openpyxl

# --- НАСТРОЙКИ ---
BOT_TOKEN = "8239333710:AAHtiMFQSNoc67WWCcHezW7m-hQ9FjxjfZ8"

# Администратор будет определен автоматически при первом запуске
# Но вы можете указать ID вручную, если знаете его
ADMIN_ID = None  # Бот сам определит

# Пути к файлам
SCHEDULE_JSON = 'schedule.json'
FILE_1 = '7d5387db9491af08bdc7e1738e02d263.xlsx'
FILE_2 = 'ae2454c52a6854c7272c28bf32f8d521.xlsx'

# Маппинг дней недели
DAYS_RU = {
    "Понедельник": 0,
    "Вторник": 1,
    "Среда": 2,
    "Четверг": 3,
    "Пятница": 4,
    "Суббота": 5,
    "Воскресенье": 6
}
DAYS_RU_REVERSE = {v: k for k, v in DAYS_RU.items()}
DAYS_SHORT = {
    "Понедельник": "ПН",
    "Вторник": "ВТ",
    "Среда": "СР",
    "Четверг": "ЧТ",
    "Пятница": "ПТ",
    "Суббота": "СБ"
}

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# --- Функции для работы с расписанием ---

def parse_schedule(file_path, week_number):
    """Парсит Excel-файл и возвращает словарь с расписанием."""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
    except Exception as e:
        logger.error(f"Ошибка при загрузке файла {file_path}: {e}")
        return {}

    schedule = {}
    groups_row = 4

    # Собираем группы
    group_columns = {}
    for col_idx, cell in enumerate(sheet[groups_row]):
        if col_idx < 2:
            continue
        group_name = cell.value
        if group_name and any(c in group_name for c in ['Э-', 'ГМУ', 'М-', 'ТД', 'ЭБ']):
            if ';' in group_name:
                for g in group_name.split(';'):
                    clean_name = g.strip()
                    group_columns[col_idx] = clean_name
                    schedule[clean_name] = {}
            else:
                group_columns[col_idx] = group_name
                schedule[group_name] = {}

    # Парсим расписание
    current_day = None
    for row_idx in range(5, sheet.max_row):
        day_cell = sheet.cell(row=row_idx, column=1).value
        time_cell = sheet.cell(row=row_idx, column=2).value

        if day_cell and day_cell in ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']:
            current_day = day_cell
            for group in schedule.keys():
                if current_day not in schedule[group]:
                    schedule[group][current_day] = {}

        if time_cell and current_day and isinstance(time_cell, str) and ':' in time_cell:
            for col_idx, group_name in group_columns.items():
                cell_value = sheet.cell(row=row_idx, column=col_idx + 1).value
                if cell_value:
                    info = str(cell_value).strip()
                    if time_cell not in schedule[group_name][current_day]:
                        schedule[group_name][current_day][time_cell] = []
                    schedule[group_name][current_day][time_cell].append(info)

    return schedule

def update_schedule():
    """Обновляет JSON-файл с расписанием из Excel-файлов."""
    try:
        logger.info("Начинаю обновление расписания...")
        
        if not os.path.exists(FILE_1) or not os.path.exists(FILE_2):
            logger.error(f"Файлы не найдены: {FILE_1}, {FILE_2}")
            return False
        
        week1 = parse_schedule(FILE_1, 1)
        week2 = parse_schedule(FILE_2, 2)
        
        if not week1 or not week2:
            logger.error("Не удалось распарсить расписание")
            return False
        
        full_schedule = {
            "1": week1,
            "2": week2
        }
        
        with open(SCHEDULE_JSON, 'w', encoding='utf-8') as f:
            json.dump(full_schedule, f, ensure_ascii=False, indent=4)
        
        logger.info("Расписание успешно обновлено!")
        return True
        
    except Exception as e:
        logger.error(f"Ошибка при обновлении расписания: {e}")
        return False

def load_schedule():
    """Загружает расписание из JSON-файла."""
    global SCHEDULE_DATA
    
    try:
        if os.path.exists(SCHEDULE_JSON):
            with open(SCHEDULE_JSON, 'r', encoding='utf-8') as f:
                SCHEDULE_DATA = json.load(f)
            logger.info("Расписание загружено из JSON")
        else:
            logger.warning("JSON-файл не найден, создаю из Excel...")
            if update_schedule():
                with open(SCHEDULE_JSON, 'r', encoding='utf-8') as f:
                    SCHEDULE_DATA = json.load(f)
            else:
                SCHEDULE_DATA = {"1": {}, "2": {}}
                logger.error("Не удалось загрузить расписание")
    except Exception as e:
        logger.error(f"Ошибка при загрузке расписания: {e}")
        SCHEDULE_DATA = {"1": {}, "2": {}}

SCHEDULE_DATA = {}

# --- Вспомогательные функции ---

def get_all_groups():
    """Собирает все уникальные группы из обеих недель."""
    groups = set()
    for week in ['1', '2']:
        for group in SCHEDULE_DATA.get(week, {}).keys():
            groups.add(group)
    return sorted(list(groups))

def get_week_number():
    """Определяет номер недели (1 или 2) на основе даты."""
    # !!! ВАЖНО: Установите реальную дату начала семестра !!!
    start_date = datetime(2024, 9, 1)
    today = datetime.now()
    delta = today - start_date
    week_num = (delta.days // 7) % 2 + 1
    return str(week_num)

def get_schedule_for_day(group_name, day_name, week_num=None):
    """Возвращает расписание для группы на конкретный день."""
    if week_num is None:
        week_num = get_week_number()
    
    if group_name not in SCHEDULE_DATA.get(week_num, {}):
        return None
    
    week_schedule = SCHEDULE_DATA[week_num][group_name]
    
    if day_name not in week_schedule:
        return f"📭 На {day_name} занятий нет."
    
    day_schedule = week_schedule[day_name]
    
    response = f"📚 *{day_name} ({week_num} неделя)*\n"
    response += f"Группа: *{group_name}*\n\n"
    
    sorted_times = sorted(day_schedule.keys())
    
    for time_slot in sorted_times:
        response += f"🕐 *{time_slot}*\n"
        for item in day_schedule[time_slot]:
            response += f"   • {item}\n"
        response += "\n"
    
    return response

def get_week_schedule(group_name):
    """Возвращает расписание на всю неделю."""
    week_num = get_week_number()
    
    if group_name not in SCHEDULE_DATA.get(week_num, {}):
        return None
    
    week_schedule = SCHEDULE_DATA[week_num][group_name]
    
    response = f"📅 *Расписание на неделю ({week_num} неделя)*\n"
    response += f"Группа: *{group_name}*\n\n"
    
    for day in ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]:
        if day in week_schedule:
            response += f"*{day}*\n"
            day_schedule = week_schedule[day]
            sorted_times = sorted(day_schedule.keys())
            for time_slot in sorted_times:
                for item in day_schedule[time_slot]:
                    response += f"   {time_slot} — {item}\n"
            response += "\n"
        else:
            response += f"*{day}* — выходной\n\n"
    
    return response

# --- Клавиатуры ---

def get_groups_keyboard():
    """Создает клавиатуру со списком групп."""
    groups = get_all_groups()
    if not groups:
        keyboard = [[InlineKeyboardButton("⏳ Расписание загружается...", callback_data="noop")]]
        return InlineKeyboardMarkup(keyboard)
    
    keyboard = []
    row = []
    for i, group in enumerate(groups):
        row.append(InlineKeyboardButton(group, callback_data=f"group_{group}"))
        if len(row) == 2:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    return InlineKeyboardMarkup(keyboard)

def get_day_keyboard(group_name):
    """Создает клавиатуру с днями недели для выбранной группы."""
    keyboard = [
        [
            InlineKeyboardButton("📖 Сегодня", callback_data=f"day_today_{group_name}"),
            InlineKeyboardButton("📖 Завтра", callback_data=f"day_tomorrow_{group_name}")
        ],
        [
            InlineKeyboardButton("📅 Вся неделя", callback_data=f"week_{group_name}")
        ],
        [
            InlineKeyboardButton("🔙 Назад к группам", callback_data="back_to_groups")
        ]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_day_buttons(group_name):
    """Создает кнопки для выбора конкретного дня."""
    keyboard = []
    row = []
    for day_ru, day_short in DAYS_SHORT.items():
        row.append(InlineKeyboardButton(day_short, callback_data=f"day_{day_ru}_{group_name}"))
        if len(row) == 3:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    keyboard.append([InlineKeyboardButton("🔙 Назад", callback_data=f"back_to_group_{group_name}")])
    return InlineKeyboardMarkup(keyboard)

# --- Обработчики ---

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start."""
    global ADMIN_ID
    
    # Если администратор еще не определен, определяем его
    if ADMIN_ID is None:
        ADMIN_ID = update.effective_user.id
        logger.info(f"Администратор установлен: {ADMIN_ID}")
        
        # Сохраняем ID в файл для будущих запусков
        with open('admin_id.txt', 'w') as f:
            f.write(str(ADMIN_ID))
        
        await update.message.reply_text(
            f"✅ Вы назначены администратором бота!\n"
            f"Ваш ID: {ADMIN_ID}\n\n"
            f"Теперь вы можете использовать команду /update_schedule для обновления расписания."
        )
    
    await update.message.reply_text(
        "👋 *Привет! Я бот расписания экономического факультета.*\n\n"
        "Выберите свою группу из списка ниже, и я покажу расписание на сегодня.",
        parse_mode="Markdown",
        reply_markup=get_groups_keyboard()
    )

async def update_schedule_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Команда для администратора: обновить расписание."""
    user_id = update.effective_user.id
    
    # Загружаем ID администратора из файла, если он есть
    admin_id = ADMIN_ID
    if admin_id is None and os.path.exists('admin_id.txt'):
        with open('admin_id.txt', 'r') as f:
            admin_id = int(f.read().strip())
    
    if user_id != admin_id:
        await update.message.reply_text("⛔ У вас нет прав для этой команды.")
        return
    
    await update.message.reply_text("🔄 Обновляю расписание...")
    
    if update_schedule():
        load_schedule()
        await update.message.reply_text(
            "✅ Расписание успешно обновлено!\n"
            "Теперь можно пользоваться актуальным расписанием.",
            reply_markup=get_groups_keyboard()
        )
    else:
        await update.message.reply_text(
            "❌ Не удалось обновить расписание.\n"
            "Проверьте, что файлы существуют и имеют правильный формат."
        )

async def handle_file_upload(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик загрузки Excel-файлов."""
    user_id = update.effective_user.id
    
    # Загружаем ID администратора из файла, если он есть
    admin_id = ADMIN_ID
    if admin_id is None and os.path.exists('admin_id.txt'):
        with open('admin_id.txt', 'r') as f:
            admin_id = int(f.read().strip())
    
    if user_id != admin_id:
        await update.message.reply_text("⛔ Вы не можете загружать файлы.")
        return
    
    document = update.message.document
    if not document:
        return
    
    file_name = document.file_name
    if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
        await update.message.reply_text("⚠️ Пожалуйста, загрузите файл Excel (.xlsx или .xls)")
        return
    
    file_path = f"temp_{file_name}"
    file = await document.get_file()
    await file.download_to_drive(file_path)
    
    # Определяем неделю
    if '1 неделя' in file_name or '1-я' in file_name or 'week_1' in file_name.lower():
        target_file = FILE_1
    elif '2 неделя' in file_name or '2-я' in file_name or 'week_2' in file_name.lower():
        target_file = FILE_2
    else:
        await update.message.reply_text(
            "📁 Не могу определить, для какой недели этот файл.\n"
            "Пожалуйста, переименуйте файл, указав '1 неделя' или '2 неделя'."
        )
        os.remove(file_path)
        return
    
    shutil.move(file_path, target_file)
    
    await update.message.reply_text(f"✅ Файл '{file_name}' сохранен как '{target_file}'")
    
    await update.message.reply_text("🔄 Обновляю расписание...")
    if update_schedule():
        load_schedule()
        await update.message.reply_text(
            "✅ Расписание успешно обновлено!",
            reply_markup=get_groups_keyboard()
        )
    else:
        await update.message.reply_text("❌ Ошибка при обновлении расписания.")

async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик нажатий на кнопки."""
    query = update.callback_query
    await query.answer()
    
    data = query.data
    
    if data == "noop":
        await query.edit_message_text(
            "⏳ Расписание еще загружается... Попробуйте позже.",
            reply_markup=get_groups_keyboard()
        )
        return
    
    if data == "back_to_groups":
        await query.edit_message_text(
            "👥 *Выберите группу:*",
            parse_mode="Markdown",
            reply_markup=get_groups_keyboard()
        )
        return
    
    if data.startswith("group_"):
        group_name = data[6:]
        context.user_data['current_group'] = group_name
        
        today_weekday = datetime.now().weekday()
        today_name = DAYS_RU_REVERSE.get(today_weekday)
        
        if today_name:
            schedule = get_schedule_for_day(group_name, today_name)
            if schedule:
                await query.edit_message_text(
                    schedule,
                    parse_mode="Markdown",
                    reply_markup=get_day_keyboard(group_name)
                )
            else:
                await query.edit_message_text(
                    f"❌ Группа '{group_name}' не найдена.",
                    parse_mode="Markdown",
                    reply_markup=get_groups_keyboard()
                )
        return
    
    if data.startswith("day_today_"):
        group_name = data[10:]
        today_weekday = datetime.now().weekday()
        today_name = DAYS_RU_REVERSE.get(today_weekday)
        schedule = get_schedule_for_day(group_name, today_name)
        if schedule:
            await query.edit_message_text(
                schedule,
                parse_mode="Markdown",
                reply_markup=get_day_keyboard(group_name)
            )
        return
    
    if data.startswith("day_tomorrow_"):
        group_name = data[13:]
        tomorrow = datetime.now() + timedelta(days=1)
        tomorrow_weekday = tomorrow.weekday()
        tomorrow_name = DAYS_RU_REVERSE.get(tomorrow_weekday)
        schedule = get_schedule_for_day(group_name, tomorrow_name)
        if schedule:
            await query.edit_message_text(
                schedule,
                parse_mode="Markdown",
                reply_markup=get_day_keyboard(group_name)
            )
        return
    
    if data.startswith("day_") and not data.startswith("day_today") and not data.startswith("day_tomorrow"):
        parts = data.split("_")
        if len(parts) >= 3:
            day_name = parts[1]
            group_name = "_".join(parts[2:])
            schedule = get_schedule_for_day(group_name, day_name)
            if schedule:
                await query.edit_message_text(
                    schedule,
                    parse_mode="Markdown",
                    reply_markup=get_day_buttons(group_name)
                )
        return
    
    if data.startswith("week_"):
        group_name = data[5:]
        schedule = get_week_schedule(group_name)
        if schedule:
            await query.edit_message_text(
                schedule,
                parse_mode="Markdown",
                reply_markup=get_day_keyboard(group_name)
            )
        return
    
    if data.startswith("back_to_group_"):
        group_name = data[14:]
        today_weekday = datetime.now().weekday()
        today_name = DAYS_RU_REVERSE.get(today_weekday)
        schedule = get_schedule_for_day(group_name, today_name)
        if schedule:
            await query.edit_message_text(
                schedule,
                parse_mode="Markdown",
                reply_markup=get_day_keyboard(group_name)
            )
        return

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик текстовых сообщений."""
    group_name = update.message.text.strip().upper()
    
    all_groups = get_all_groups()
    if group_name not in all_groups:
        await update.message.reply_text(
            f"❌ Группа '{group_name}' не найдена.\n"
            "Пожалуйста, выберите группу из списка ниже:",
            reply_markup=get_groups_keyboard()
        )
        return
    
    today_weekday = datetime.now().weekday()
    today_name = DAYS_RU_REVERSE.get(today_weekday)
    schedule = get_schedule_for_day(group_name, today_name)
    
    if schedule:
        await update.message.reply_text(
            schedule,
            parse_mode="Markdown",
            reply_markup=get_day_keyboard(group_name)
        )

async def show_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Команда для показа ID пользователя."""
    user_id = update.effective_user.id
    username = update.effective_user.username or "нет username"
    
    await update.message.reply_text(
        f"🆔 *Ваш ID в Telegram:*\n"
        f"`{user_id}`\n\n"
        f"👤 Username: @{username}\n\n"
        f"Если вы администратор, запишите этот ID и укажите его в коде бота."
    )

# --- Основная функция ---

def main():
    """Запуск бота."""
    global ADMIN_ID
    
    # Пытаемся загрузить ID администратора из файла
    if os.path.exists('admin_id.txt'):
        try:
            with open('admin_id.txt', 'r') as f:
                ADMIN_ID = int(f.read().strip())
            logger.info(f"Загружен ID администратора: {ADMIN_ID}")
        except:
            pass
    
    # Загружаем расписание
    load_schedule()
    
    application = Application.builder().token(BOT_TOKEN).build()
    
    # Регистрируем обработчики
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("update_schedule", update_schedule_command))
    application.add_handler(CommandHandler("myid", show_id))  # Новая команда для получения ID
    application.add_handler(CallbackQueryHandler(button_callback))
    application.add_handler(MessageHandler(filters.Document.ALL, handle_file_upload))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    
    print("🤖 Бот запущен и готов к работе!")
    if ADMIN_ID:
        print(f"👤 Администратор ID: {ADMIN_ID}")
    else:
        print("👤 Администратор не назначен. Первый запуск /start назначит администратора.")
    print("📅 Расписание загружено")
    print(f"📊 Найдено групп: {len(get_all_groups())}")
    print("\nДоступные команды:")
    print("  /start - Начать работу")
    print("  /myid - Показать ваш Telegram ID")
    print("  /update_schedule - Обновить расписание (только для администратора)")
    
    application.run_polling()

if __name__ == '__main__':
    main()